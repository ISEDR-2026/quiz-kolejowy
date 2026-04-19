import streamlit as st
from openpyxl import load_workbook
import random
import os
import time

st.set_page_config(layout="wide")

# ===== STYLE =====
st.markdown("""
<style>
.card {background:#1c1f26;padding:20px;border-radius:15px;margin-bottom:20px;}

.question-number {
    font-size:14px;
    color:gray;
    margin-bottom:5px;
}

.question-text {
    font-size:20px;
    font-weight:bold;
}

div[role="radiogroup"] > label {
    font-size:18px !important;
    padding:10px !important;
}

button[kind="secondary"], button[kind="primary"] {
    font-size:18px !important;
    padding:10px 20px !important;
    border-radius:10px !important;
}

.answer {padding:10px;border-radius:10px;margin:5px 0;background:#2a2e38;}
.correct {background:#2ecc71 !important;color:white;}
.wrong {background:#e74c3c !important;color:white;}

.login-box {
    background:#1c1f26;
    padding:40px;
    border-radius:15px;
    width:400px;
    margin:auto;
    margin-top:10%;
    text-align:center;
}

/* ===== NOWE: RESPONSYWNE OBRAZKI ===== */
img {
    width: auto;
    height: auto;
    max-width: 220px;
}

.img-big img {
    max-width: 500px;
}
</style>
""", unsafe_allow_html=True)

st.markdown("<script>window.scrollTo(0,0);</script>", unsafe_allow_html=True)

HASLO = "szczecin26"

BIG_IMAGES = {1157,1158,1159,1160,1161,1162,1163,1164,1165,1166,1169,1170,1171}

def is_big_image(nr):
    return nr in BIG_IMAGES

def norm(txt):
    if not txt:
        return ""
    return " ".join(str(txt).lower().strip().split())

IR1_FULL = norm("Ir - 1 Instrukcja o prowadzeniu ruchu pociągów")
IE1_FULL = norm("Ie - 1 Instrukcja sygnalizacji")

# ===== LOGOWANIE =====
if "auth" not in st.session_state:
    st.session_state.auth = False

if not st.session_state.auth:
    st.markdown("""
    <div class="login-box">
        <h2>🔐 Dostęp</h2>
        <p style="color:gray;">
        Baza pytań sprawdzających wiedzę na stanowisku Dyżurny Ruchu
        </p>
    """, unsafe_allow_html=True)

    password = st.text_input("Hasło", type="password")

    if st.button("Wejdź"):
        if password == HASLO:
            st.session_state.auth = True
            st.rerun()
        else:
            st.error("❌ Niepoprawne hasło")

    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()

# ===== WCZYTANIE =====
wb = load_workbook("quiz.xlsx")
ws = wb.active

questions = []
for row in ws.iter_rows(min_row=5):
    correct_raw = row[5].value
    if correct_raw is None:
        continue

    questions.append({
        "nr": int(row[0].value),
        "q": row[1].value,
        "answers": {"a": row[2].value, "b": row[3].value, "c": row[4].value},
        "correct": str(correct_raw).lower().strip(),
        "img": row[6].value if len(row) > 6 else None,
        "section": norm(row[7].value if len(row) > 7 else "")
    })

def filter_sections(q_list, selected):
    if selected == "Wszystkie":
        return q_list
    if selected == "Instrukcja Ir1":
        return [q for q in q_list if q["section"] == IR1_FULL]
    if selected == "Sygnalizacja":
        return [q for q in q_list if q["section"] == IE1_FULL]
    if selected == "Inne":
        return [q for q in q_list if q["section"] not in [IR1_FULL, IE1_FULL]]

# ===== SESSION =====
if "index" not in st.session_state:
    st.session_state.update({
        "index": 0,
        "score": 0,
        "started": False,
        "selected": [],
        "answers_log": [],
        "mode": "learn",
        "answered": False,
        "last_choice": None
    })

# ===== START =====
if not st.session_state.started:

    st.markdown("""
    <h2 style='text-align:center;'>
    🚆 Baza pytań sprawdzających wiedzę na stanowisku Dyżurny Ruchu<br>
    <span style='font-size:16px;color:gray;'>Tryb nauki i egzaminu</span>
    </h2>
    """, unsafe_allow_html=True)

    mode = st.radio("Tryb:", ["Nauka", "Egzamin"])

    if mode == "Nauka":
        option = st.selectbox("Ilość pytań:", ["10", "25", "Od pytania"])
        section = st.radio("Dział:", ["Wszystkie", "Instrukcja Ir1", "Sygnalizacja", "Inne"])
        start_q = st.number_input("Od pytania:", 1, len(questions), 1) if option == "Od pytania" else 1

    if st.button("Start"):

        if mode == "Nauka":
            filtered = filter_sections(questions, section)

            if option == "10":
                selected = random.sample(filtered, min(10, len(filtered)))
            elif option == "25":
                selected = random.sample(filtered, min(25, len(filtered)))
            else:
                selected = [q for q in filtered if q["nr"] >= start_q]

        else:
            signal_q = [q for q in questions if q["section"] == IE1_FULL]
            other_q = [q for q in questions if q["section"] != IE1_FULL]

            selected_signal = random.sample(signal_q, min(5, len(signal_q)))
            selected_other = random.sample(other_q, 30 - len(selected_signal))

            selected = selected_signal + selected_other
            random.shuffle(selected)

        st.session_state.update({
            "selected": selected,
            "started": True,
            "index": 0,
            "score": 0,
            "answers_log": [],
            "mode": "learn" if mode == "Nauka" else "exam",
            "answered": False
        })
        st.rerun()

# ===== QUIZ =====
else:
    q_list = st.session_state.selected
    i = st.session_state.index
    finished = i >= len(q_list)

    st.progress(i / len(q_list))

    if finished:

        total = len(q_list)
        correct = st.session_state.score
        percent = int((correct / total) * 100)

        st.markdown(f"## Wynik: {correct}/{total} ({percent}%)")
        st.markdown("---")

        for idx, item in enumerate(st.session_state.answers_log, start=1):
            st.markdown(f"<div class='question-number'>Pytanie {idx} (nr {item['nr']})</div>", unsafe_allow_html=True)
            st.markdown(f"<div class='question-text'>{item['q']}</div>", unsafe_allow_html=True)

            if item["img"] == "img":
                path = f"image/{item['nr']}.png"
                if os.path.exists(path):
                    if is_big_image(item['nr']):
                        st.markdown("<div class='img-big'>", unsafe_allow_html=True)
                        st.image(path)
                        st.markdown("</div>", unsafe_allow_html=True)
                    else:
                        st.image(path)

            for k, txt in item["answers"].items():
                if k == item["correct_answer"]:
                    st.markdown(f"<div class='answer correct'>• {txt}</div>", unsafe_allow_html=True)
                elif k == item["selected"]:
                    st.markdown(f"<div class='answer wrong'>• {txt}</div>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<div class='answer'>• {txt}</div>", unsafe_allow_html=True)

        if st.button("Restart"):
            st.session_state.update({
                "index": 0,
                "score": 0,
                "started": False,
                "selected": [],
                "answers_log": [],
                "answered": False
            })
            st.rerun()

    else:
        q = q_list[i]

        st.markdown(f"""
        <div class='card'>
            <div class='question-number'>Pytanie {i+1} (nr {q['nr']})</div>
            <div class='question-text'>{q['q']}</div>
        </div>
        """, unsafe_allow_html=True)

        if q["img"] == "img":
            path = f"image/{q['nr']}.png"
            if os.path.exists(path):
                if is_big_image(q['nr']):
                    st.markdown("<div class='img-big'>", unsafe_allow_html=True)
                    st.image(path)
                    st.markdown("</div>", unsafe_allow_html=True)
                else:
                    st.image(path)

        if f"shuffled_{i}" not in st.session_state:
            items = list(q["answers"].items())
            random.shuffle(items)
            st.session_state[f"shuffled_{i}"] = items

        shuffled = st.session_state[f"shuffled_{i}"]
        options = [k for k, _ in shuffled]

        choice = st.radio(
            "Odpowiedzi:",
            options,
            format_func=lambda x: f"{dict(shuffled)[x]}",
            key=f"q_{i}"
        )

        if st.session_state.mode == "learn":

            if not st.session_state.answered:
                if st.button("Zatwierdź"):
                    is_correct = choice == q["correct"]

                    if is_correct:
                        st.session_state.score += 1

                    st.session_state.answers_log.append({
                        "nr": q["nr"],
                        "q": q["q"],
                        "answers": dict(shuffled),
                        "selected": choice,
                        "correct_answer": q["correct"],
                        "correct": is_correct,
                        "section": q["section"],
                        "img": q["img"]
                    })

                    st.session_state.last_choice = choice
                    st.session_state.answered = True
                    st.rerun()

            else:
                for k, txt in dict(shuffled).items():
                    cls = "answer"
                    if k == q["correct"]:
                        cls += " correct"
                    elif k == st.session_state.last_choice:
                        cls += " wrong"

                    st.markdown(f"<div class='{cls}'>{txt}</div>", unsafe_allow_html=True)

                if st.button("➡ Dalej"):
                    st.session_state.index += 1
                    st.session_state.answered = False
                    st.rerun()

        else:
            if st.button("Zatwierdź"):
                is_correct = choice == q["correct"]

                if is_correct:
                    st.session_state.score += 1

                st.session_state.answers_log.append({
                    "nr": q["nr"],
                    "q": q["q"],
                    "answers": dict(shuffled),
                    "selected": choice,
                    "correct_answer": q["correct"],
                    "correct": is_correct,
                    "section": q["section"],
                    "img": q["img"]
                })

                st.session_state.index += 1
                st.rerun()
