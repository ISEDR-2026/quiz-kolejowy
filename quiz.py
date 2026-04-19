import tkinter as tk
from openpyxl import load_workbook
import random

# ===== Wczytanie Excela =====
wb = load_workbook("quiz.xlsx")
ws = wb.active

questions = []

for row in ws.iter_rows(min_row=5):
    nr = row[0].value
    q = row[1].value
    a = row[2]
    b = row[3]
    c = row[4]

    answers = [
        ("a", a.value, a.font.color),
        ("b", b.value, b.font.color),
        ("c", c.value, c.font.color),
    ]

    correct = None

    for label, text, color in answers:
        if color and color.rgb:
            if "00FF00" in str(color.rgb) or "008000" in str(color.rgb):
                correct = label

    questions.append({
        "nr": nr,
        "q": q,
        "answers": answers,
        "correct": correct
    })

random.shuffle(questions)

# ===== GUI =====
root = tk.Tk()
root.title("Quiz kolejowy")
root.geometry("700x400")

current = 0
score = 0

# ===== Funkcja wyświetlania pytania =====
def show_question():
    global current

    if current >= 10:
        question_label.config(text=f"Koniec! Wynik: {score}/10")
        for btn in buttons:
            btn.config(state="disabled")
        return

    q = questions[current]

    question_label.config(text=f"Pytanie {q['nr']}: {q['q']}")

    for i, (label, text, _) in enumerate(q["answers"]):
        buttons[i].config(text=f"{label}) {text}", state="normal")

# ===== Sprawdzanie odpowiedzi =====
def check_answer(choice):
    global current, score

    if choice == questions[current]["correct"]:
        score += 1

    current += 1
    show_question()

# ===== UI =====
question_label = tk.Label(root, text="", wraplength=650, font=("Arial", 12))
question_label.pack(pady=20)

buttons = []
for i in range(3):
    btn = tk.Button(root, text="", width=80, command=lambda i=i: check_answer(["a","b","c"][i]))
    btn.pack(pady=5)
    buttons.append(btn)

show_question()

root.mainloop()