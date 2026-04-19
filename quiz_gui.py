import tkinter as tk
from openpyxl import load_workbook
import random

# ===== Wczytanie Excela =====
wb = load_workbook("quiz.xlsx")
ws = wb.active

dzial = ws["C3"].value

all_questions = []

for row in ws.iter_rows(min_row=5):
    nr = row[0].value
    q = row[1].value
    a = row[2].value
    b = row[3].value
    c = row[4].value
    correct = row[5].value

    if not q or not correct:
        continue

    all_questions.append({
        "nr": nr,
        "q": q,
        "answers": [("a", a), ("b", b), ("c", c)],
        "correct": correct.lower()
    })

# ===== GUI =====
root = tk.Tk()
root.title("Quiz kolejowy")
root.geometry("900x650")

current = 0
score = 0
questions = []
mode = ""
time_left = 0

# ===== TIMER =====
def update_timer():
    global time_left

    if mode != "exam":
        return

    if time_left > 0:
        mins = time_left // 60
        secs = time_left % 60
        timer_label.config(text=f"Czas: {mins:02}:{secs:02}")
        time_left -= 1
        root.after(1000, update_timer)
    else:
        end_exam()

# ===== START QUIZU =====
def start_quiz(count):
    global questions, current, score, time_left

    random.shuffle(all_questions)

    if count == "all":
        questions = all_questions.copy()
    else:
        questions = all_questions[:int(count)]

    current = 0
    score = 0

    if mode == "exam":
        time_left = 1800  # 30 minut
        timer_label.pack(pady=5)
        update_timer()
    else:
        timer_label.pack_forget()

    learn_frame.pack_forget()
    quiz_frame.pack()

    show_question()

# ===== WYŚWIETLANIE =====
def show_question():
    global current

    if current >= len(questions):
        end_exam()
        return

    q = questions[current]

    dzial_label.config(text=f"[{dzial}]")
    nr_label.config(text=f"Pytanie {q['nr']}")
    question_label.config(text=q["q"])
    feedback_label.config(text="")

    for i, (label, text) in enumerate(q["answers"]):
        buttons[i].config(
            text=f"{label}) {text}",
            bg="SystemButtonFace",
            state="normal",
            wraplength=850,
            justify="left"
        )

# ===== SPRAWDZANIE =====
def check_answer(choice):
    global current, score

    correct = questions[current]["correct"]

    if choice == correct:
        score += 1

    if mode == "learn":
        for i, (label, _) in enumerate(questions[current]["answers"]):
            if label == correct:
                buttons[i].config(bg="lightgreen")
            elif label == choice:
                buttons[i].config(bg="tomato")

        if choice == correct:
            feedback_label.config(text="✔ DOBRZE", fg="green")
        else:
            feedback_label.config(text=f"❌ ŹLE (poprawna: {correct})", fg="red")

        next_btn.pack(pady=10)
    else:
        current += 1
        show_question()

# ===== NASTĘPNE =====
def next_question():
    global current
    current += 1
    next_btn.pack_forget()
    show_question()

# ===== KONIEC =====
def end_exam():
    percent = int((score / len(questions)) * 100)
    wynik = "ZALICZONE ✅" if percent >= 80 else "NIEZALICZONE ❌"

    question_label.config(
        text=f"KONIEC TESTU\n\nWynik: {score}/{len(questions)} ({percent}%)\n{wynik}"
    )

    for btn in buttons:
        btn.pack_forget()

    next_btn.pack_forget()
    timer_label.pack_forget()

# ===== WYBÓR TRYBU =====
def choose_learn():
    global mode
    mode = "learn"
    start_frame.pack_forget()
    learn_frame.pack()

def choose_exam():
    global mode
    mode = "exam"

    start_frame.pack_forget()

    exam_info = tk.Label(
        root,
        text=f"Tryb egzamin\n\nLiczba pytań: {len(all_questions)}\nCzas: 30 minut",
        font=("Arial", 14)
    )
    exam_info.pack(pady=50)

    tk.Button(root, text="Rozpocznij egzamin", command=lambda: start_quiz("all")).pack()

# ===== START SCREEN =====
start_frame = tk.Frame(root)
start_frame.pack(pady=40)

tk.Label(
    start_frame,
    text=f"Wybierz dział:\n{dzial}",
    font=("Arial", 14)
).pack(pady=10)

tk.Button(start_frame, text="NAUKA", width=20, command=choose_learn).pack(pady=5)
tk.Button(start_frame, text="EGZAMIN", width=20, command=choose_exam).pack(pady=5)

# ===== WYBÓR PYTAŃ (NAUKA) =====
learn_frame = tk.Frame(root)

tk.Label(learn_frame, text="Wybierz liczbę pytań:", font=("Arial", 14)).pack(pady=10)

tk.Button(learn_frame, text="10", width=20, command=lambda: start_quiz(10)).pack(pady=5)
tk.Button(learn_frame, text="25", width=20, command=lambda: start_quiz(25)).pack(pady=5)
tk.Button(learn_frame, text="Wszystkie", width=20, command=lambda: start_quiz("all")).pack(pady=5)

entry = tk.Entry(learn_frame)
entry.pack(pady=5)

tk.Button(learn_frame, text="Start (własna liczba)",
          command=lambda: start_quiz(entry.get())).pack(pady=5)

# ===== QUIZ =====
quiz_frame = tk.Frame(root)

timer_label = tk.Label(quiz_frame, text="", font=("Arial", 12), fg="red")

dzial_label = tk.Label(quiz_frame, text="", font=("Arial", 12))
dzial_label.pack(pady=5)

nr_label = tk.Label(quiz_frame, text="", font=("Arial", 12))
nr_label.pack(pady=5)

question_label = tk.Label(
    quiz_frame,
    text="",
    wraplength=850,
    font=("Arial", 14, "bold"),
    justify="center"
)
question_label.pack(pady=20)

buttons = []
for i in range(3):
    btn = tk.Button(
        quiz_frame,
        text="",
        width=120,
        anchor="w",
        justify="left",
        command=lambda i=i: check_answer(["a","b","c"][i])
    )
    btn.pack(pady=5)
    buttons.append(btn)

feedback_label = tk.Label(quiz_frame, text="", font=("Arial", 12))
feedback_label.pack(pady=10)

next_btn = tk.Button(quiz_frame, text="Następne pytanie", command=next_question)

root.mainloop()